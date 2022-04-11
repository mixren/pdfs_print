import os
import time

def check_server_connection(ip: str):
    import socket
    try:
        socket.gethostbyaddr(ip)
        return True
    except socket.herror:
        False

def get_dirs_from_excel(file_excel: str, char_column: str) -> list:
    from openpyxl import load_workbook
    wb = load_workbook(filename=file_excel)
    sheet = wb.active
    col = sheet[char_column]
    return [".".join(entry.value.strip().split(".")[0:2]) for entry in col]

def get_drawings_from_excel(file_excel: str, char_column: str) -> list:
    from openpyxl import load_workbook
    wb = load_workbook(filename=file_excel)
    sheet = wb.active
    col = sheet[char_column]
    return [".".join(entry.value.strip().split(".")[:-1]) for entry in col]

def extract_isometry(drawing: str)-> str:
    return drawing.strip().rsplit(".", 1)[0]

def modify_last_dot_to_dash(drawing:str):
    return "-".join(drawing.rsplit(".", 1))
    
def modify_remove_cwt(drawing:str):
    return drawing.split(".", 1)[1]

def get_existing_dir_path(path: str, dir: str)->str:
    dir_path = os.path.join(path, dir)
    if not os.path.exists(dir_path):
        try:
            dir_path = os.path.join(path, find_similar_name_dir(path, dir))
        except:
            raise
    return dir_path

def find_similar_name_dir(path: str, dir: str)->str:
    try:
        for _, dirs, _ in os.walk(path):
            return [d for d in dirs if dir in d][0]
    except:
        raise
    
def find_paths_for_rasejumi_pdfs_2(path: str, dir_names: list)->list:
    return_paths = []
    for dir in dir_names:
        dir_path=""
        try:
            dir_path = get_existing_dir_path(path, dir)  # Throwable
        except:
            print(f"No directory named '{dir}'")
            continue
        rasejumi="Rasejumi ceham"
        try:
            rasejumi_path = get_existing_dir_path(dir_path, rasejumi)  # Throwable
            for root, _, files in os.walk(rasejumi_path):
                return_paths.extend([os.path.join(root, file) for file in files])
        except:
            print(f"No directory named '{rasejumi}' in {dir}")
            continue
    return return_paths

def find_paths_for_rasejumi_pdfs(path: str, drawings: list):
    return_path=[]
    for drawing in drawings:
        isometry = extract_isometry(drawing)
        dir_path=""
        try:
            dir_path = get_existing_dir_path(path, isometry)  # Throwable
        except:
            print(f"No directory named '{isometry}'")
            continue
        rasejumi="Rasejumi ceham"
        rasejumi_path=""
        try:
            rasejumi_path = get_existing_dir_path(dir_path, rasejumi)  # Throwable
        except:
            print(f"No directory named '{rasejumi}' in {dir}")
            continue
        dr_mod_1=f'{modify_remove_cwt(drawing)}'
        dr_mod_2=f'{modify_last_dot_to_dash(modify_remove_cwt(drawing))}'
        for root, _, files in os.walk(rasejumi_path):
            found=False
            for file in files:
                if dr_mod_1 in file or dr_mod_2 in file:
                    return_path.append(os.path.join(root, file))
                    found=True
                    break
            if not found:
                print(f"No file '{drawing}' is not found.")

        '''if os.path.exists(path_1):
            return_path.append(path_1)
        elif os.path.exists(path_2):
            return_path.append(path_2)
        else:
            print(f"No file '{drawing}' nor '{modify(drawing)}' is found.")'''
    return return_path

def merge_pdfs(paths: list):
    from pikepdf import Pdf
    pdf = Pdf.new()
    for path in paths:
        src = Pdf.open(path)
        pdf.pages.append(src.pages[0])
        src.close()
    pdf.save('MergedPdf.pdf')
    pdf.close()

def main():
    CWT_server_ip = "192.168.4.11"
    if not check_server_connection(CWT_server_ip):
        print("You are not connected to the CWT server")
        return

    root_2021 = "Z:\CWT Documents\\2021"
    dir_21_81 = "MP21-81 PP108_POX_CS"
    dir_21_80 = "MP21-80 BP101_SL_CS"
    file = "print_3.xlsx"
    root_path=os.path.join(root_2021, dir_21_80)

    drawings = get_drawings_from_excel(file_excel=file, char_column="A")
    paths_ras = find_paths_for_rasejumi_pdfs(root_path, drawings)
    if len(drawings) == len(paths_ras):
        print("All files found.")
    #dirs_from_excel = get_dirs_from_excel(file_excel=file, char_column="A")
    #paths_rasejumi = find_paths_for_rasejumi_pdfs_2(root_path, dirs_from_excel)
    merge_pdfs(paths_ras)

start = time.time()
main()
end = time.time()
print(f"Success \nTime elapsed: {round(end-start,4)} sec")
